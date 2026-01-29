import sqlite3
from datetime import datetime, date, timedelta
from typing import List, Optional, Tuple
import sys
import shutil

try:
    from PyQt5.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QTabWidget, QLabel, QLineEdit, QPushButton, QTableWidget,
        QTableWidgetItem, QMessageBox, QDialog, QTextEdit, QGroupBox,
        QFormLayout, QHeaderView, QFileDialog, QDateEdit, QGridLayout
    )
    from PyQt5.QtCore import Qt, QDate
    from PyQt5.QtGui import QFont
except ImportError:
    print("ERRO: PyQt5 não está instalado!")
    print("Instale com: pip install PyQt5")
    sys.exit(1)


class GerenciadorDeposito:
    def __init__(self, db_name: str = "deposito.db"):
        self.db_name = db_name
        self.criar_tabelas()
    
    def conectar(self) -> sqlite3.Connection:
        return sqlite3.connect(self.db_name)
    
    def criar_tabelas(self):
        conn = self.conectar()
        cursor = conn.cursor()
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS produtos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL,
                descricao TEXT,
                categoria TEXT,
                quantidade INTEGER NOT NULL DEFAULT 0,
                localizacao TEXT,
                data_cadastro TEXT NOT NULL
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS movimentacoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                produto_id INTEGER NOT NULL,
                tipo TEXT NOT NULL,
                quantidade INTEGER NOT NULL,
                data_movimentacao TEXT NOT NULL,
                observacao TEXT,
                FOREIGN KEY (produto_id) REFERENCES produtos(id)
            )
        """)
        
        conn.commit()
        conn.close()
    
    def adicionar_produto(self, nome: str, quantidade: int = 0, 
                         descricao: str = "", categoria: str = "", 
                         localizacao: str = "") -> int:
        conn = self.conectar()
        cursor = conn.cursor()
        
        data_atual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        cursor.execute("""
            INSERT INTO produtos (nome, descricao, categoria, quantidade, localizacao, data_cadastro)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (nome.upper(), descricao.upper(), categoria.upper(), quantidade, localizacao.upper(), data_atual))
        
        produto_id = cursor.lastrowid
        
        if quantidade > 0:
            cursor.execute("""
                INSERT INTO movimentacoes (produto_id, tipo, quantidade, data_movimentacao, observacao)
                VALUES (?, 'ENTRADA', ?, ?, 'Estoque inicial')
            """, (produto_id, quantidade, data_atual))
        
        conn.commit()
        conn.close()
        
        return produto_id
    
    def listar_produtos(self, categoria: Optional[str] = None) -> List[Tuple]:
        conn = self.conectar()
        cursor = conn.cursor()
        
        if categoria:
            cursor.execute("""
                SELECT id, nome, descricao, categoria, quantidade, localizacao
                FROM produtos WHERE categoria = ?
                ORDER BY nome
            """, (categoria,))
        else:
            cursor.execute("""
                SELECT id, nome, descricao, categoria, quantidade, localizacao
                FROM produtos ORDER BY nome
            """)
        
        produtos = cursor.fetchall()
        conn.close()
        
        return produtos
    
    def buscar_produto(self, produto_id: int) -> Optional[Tuple]:
        conn = self.conectar()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, nome, descricao, categoria, quantidade, localizacao, data_cadastro
            FROM produtos WHERE id = ?
        """, (produto_id,))
        
        produto = cursor.fetchone()
        conn.close()
        
        return produto
    
    def buscar_produto_por_nome(self, nome: str) -> List[Tuple]:
        conn = self.conectar()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, nome, descricao, categoria, quantidade, localizacao
            FROM produtos WHERE nome LIKE ?
            ORDER BY nome
        """, (f"%{nome}%",))
        
        produtos = cursor.fetchall()
        conn.close()
        
        return produtos
    
    def atualizar_produto(self, produto_id: int, **kwargs):
        campos_permitidos = ['nome', 'descricao', 'categoria', 'localizacao']
        campos = []
        valores = []
        
        for campo, valor in kwargs.items():
            if campo in campos_permitidos:
                campos.append(f"{campo} = ?")
                valores.append(valor)
        
        if not campos:
            return False
        
        valores.append(produto_id)
        
        conn = self.conectar()
        cursor = conn.cursor()
        
        query = f"UPDATE produtos SET {', '.join(campos)} WHERE id = ?"
        cursor.execute(query, valores)
        
        conn.commit()
        linhas_afetadas = cursor.rowcount
        conn.close()
        
        return linhas_afetadas > 0
    
    def registrar_entrada(self, produto_id: int, quantidade: int, 
                         observacao: str = "") -> bool:
        conn = self.conectar()
        cursor = conn.cursor()
        
        cursor.execute("SELECT quantidade FROM produtos WHERE id = ?", (produto_id,))
        resultado = cursor.fetchone()
        
        if not resultado:
            conn.close()
            return False
        
        quantidade_atual = resultado[0]
        nova_quantidade = quantidade_atual + quantidade
        
        cursor.execute("UPDATE produtos SET quantidade = ? WHERE id = ?", (nova_quantidade, produto_id))
        
        data_atual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute("""
            INSERT INTO movimentacoes (produto_id, tipo, quantidade, data_movimentacao, observacao)
            VALUES (?, 'ENTRADA', ?, ?, ?)
        """, (produto_id, quantidade, data_atual, observacao))
        
        conn.commit()
        conn.close()
        
        return True
    
    def registrar_saida(self, produto_id: int, quantidade: int, 
                       observacao: str = "") -> bool:
        conn = self.conectar()
        cursor = conn.cursor()
        
        cursor.execute("SELECT quantidade FROM produtos WHERE id = ?", (produto_id,))
        resultado = cursor.fetchone()
        
        if not resultado:
            conn.close()
            return False
        
        quantidade_atual = resultado[0]
        
        if quantidade_atual < quantidade:
            conn.close()
            return False
        
        nova_quantidade = quantidade_atual - quantidade
        
        cursor.execute("UPDATE produtos SET quantidade = ? WHERE id = ?", (nova_quantidade, produto_id))
        
        data_atual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute("""
            INSERT INTO movimentacoes (produto_id, tipo, quantidade, data_movimentacao, observacao)
            VALUES (?, 'SAIDA', ?, ?, ?)
        """, (produto_id, quantidade, data_atual, observacao))
        
        conn.commit()
        conn.close()
        
        return True
    
    def listar_movimentacoes(self, produto_id: Optional[int] = None, 
                            data_inicio: Optional[str] = None, 
                            data_fim: Optional[str] = None) -> List[Tuple]:
        conn = self.conectar()
        cursor = conn.cursor()
        
        query = """
            SELECT m.id, p.nome, m.tipo, m.quantidade, m.data_movimentacao, m.observacao
            FROM movimentacoes m
            JOIN produtos p ON m.produto_id = p.id
            WHERE 1=1
        """
        params = []
        
        if produto_id:
            query += " AND m.produto_id = ?"
            params.append(produto_id)
        
        if data_inicio:
            query += " AND DATE(m.data_movimentacao) >= ?"
            params.append(data_inicio)
        
        if data_fim:
            query += " AND DATE(m.data_movimentacao) <= ?"
            params.append(data_fim)
        
        query += " ORDER BY m.data_movimentacao DESC LIMIT 500"
        
        cursor.execute(query, params)
        movimentacoes = cursor.fetchall()
        conn.close()
        
        return movimentacoes
    
    def produtos_estoque_baixo(self, limite: int = 10) -> List[Tuple]:
        conn = self.conectar()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, nome, categoria, quantidade, localizacao
            FROM produtos
            WHERE quantidade <= ?
            ORDER BY quantidade ASC
        """, (limite,))
        
        produtos = cursor.fetchall()
        conn.close()
        
        return produtos
    
    def relatorio_estoque(self) -> dict:
        conn = self.conectar()
        cursor = conn.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM produtos")
        total_produtos = cursor.fetchone()[0]
        
        cursor.execute("SELECT SUM(quantidade) FROM produtos")
        total_itens = cursor.fetchone()[0] or 0
        
        cursor.execute("""
            SELECT categoria, COUNT(*), SUM(quantidade)
            FROM produtos
            GROUP BY categoria
        """)
        por_categoria = cursor.fetchall()
        
        conn.close()
        
        return {
            'total_produtos': total_produtos,
            'total_itens': total_itens,
            'por_categoria': por_categoria
        }


class DialogMovimentacao(QDialog):
    def __init__(self, parent, produto_id, produto_nome, tipo, qtd_atual=None):
        super().__init__(parent)
        self.produto_id = produto_id
        self.tipo = tipo
        self.deposito = parent.deposito
        self.parent = parent
        
        self.setWindowTitle(f"{tipo} - {produto_nome}")
        self.setModal(True)
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout()
        
        titulo = QLabel(f"Produto: {produto_nome}")
        titulo.setFont(QFont("Arial", 12, QFont.Bold))
        layout.addWidget(titulo)
        
        if qtd_atual is not None:
            qtd_label = QLabel(f"Quantidade atual: {qtd_atual}")
            layout.addWidget(qtd_label)
        
        form_layout = QFormLayout()
        
        self.qtd_input = QLineEdit()
        form_layout.addRow("Quantidade:", self.qtd_input)
        
        self.obs_input = QLineEdit()
        form_layout.addRow("Observação:", self.obs_input)
        
        layout.addLayout(form_layout)
        
        btn_confirmar = QPushButton("Confirmar")
        btn_confirmar.clicked.connect(self.confirmar)
        layout.addWidget(btn_confirmar)
        
        self.setLayout(layout)
    
    def confirmar(self):
        try:
            quantidade = int(self.qtd_input.text())
            observacao = self.obs_input.text().strip()
            
            if self.tipo == "ENTRADA":
                sucesso = self.deposito.registrar_entrada(self.produto_id, quantidade, observacao)
            else:
                sucesso = self.deposito.registrar_saida(self.produto_id, quantidade, observacao)
            
            if sucesso:
                QMessageBox.information(self, "Sucesso", f"{self.tipo} registrada com sucesso!")
                self.parent.atualizar_lista_produtos()
                self.parent.atualizar_movimentacoes()
                self.accept()
            else:
                msg = "Quantidade insuficiente em estoque!" if self.tipo == "SAÍDA" else "Erro ao registrar!"
                QMessageBox.critical(self, "Erro", msg)
        
        except ValueError:
            QMessageBox.critical(self, "Erro", "Digite uma quantidade válida!")


class DialogEditarProduto(QDialog):
    def __init__(self, parent, produto_id):
        super().__init__(parent)
        self.produto_id = produto_id
        self.deposito = parent.deposito
        self.parent = parent
        
        self.setWindowTitle("Editar Produto")
        self.setModal(True)
        self.setMinimumWidth(500)
        
        produto = self.deposito.buscar_produto(produto_id)
        if not produto:
            QMessageBox.critical(self, "Erro", "Produto não encontrado!")
            self.reject()
            return
        
        layout = QVBoxLayout()
        
        titulo = QLabel("Editar Cadastro de Produto")
        titulo.setFont(QFont("Arial", 14, QFont.Bold))
        titulo.setAlignment(Qt.AlignCenter)
        layout.addWidget(titulo)
        
        id_label = QLabel(f"ID: {produto[0]}")
        id_label.setFont(QFont("Arial", 10))
        layout.addWidget(id_label)
        
        form_layout = QFormLayout()
        
        self.nome_input = QLineEdit()
        self.nome_input.setText(produto[1])
        form_layout.addRow("Nome:", self.nome_input)
        
        self.categoria_input = QLineEdit()
        self.categoria_input.setText(produto[3] or "")
        form_layout.addRow("Categoria:", self.categoria_input)
        
        self.localizacao_input = QLineEdit()
        self.localizacao_input.setText(produto[5] or "")
        form_layout.addRow("Localização:", self.localizacao_input)
        
        self.descricao_input = QLineEdit()
        self.descricao_input.setText(produto[2] or "")
        form_layout.addRow("Descrição:", self.descricao_input)
        
        qtd_info = QLabel(f"Quantidade atual: {produto[4]}")
        qtd_info.setStyleSheet("color: #0066cc; font-weight: bold;")
        form_layout.addRow("", qtd_info)
        
        obs_label = QLabel("(Use Entrada/Saída para alterar quantidade)")
        obs_label.setStyleSheet("color: #666; font-size: 9pt; font-style: italic;")
        form_layout.addRow("", obs_label)
        
        layout.addLayout(form_layout)
        
        btn_layout = QHBoxLayout()
        
        btn_salvar = QPushButton("Salvar")
        btn_salvar.clicked.connect(self.salvar)
        btn_layout.addWidget(btn_salvar)
        
        btn_cancelar = QPushButton("Cancelar")
        btn_cancelar.clicked.connect(self.reject)
        btn_layout.addWidget(btn_cancelar)
        
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
    
    def salvar(self):
        nome = self.nome_input.text().strip()
        categoria = self.categoria_input.text().strip()
        localizacao = self.localizacao_input.text().strip()
        descricao = self.descricao_input.text().strip()
        
        if not nome:
            QMessageBox.critical(self, "Erro", "Nome do produto é obrigatório!")
            return
        
        sucesso = self.deposito.atualizar_produto(
            self.produto_id,
            nome=nome.upper(),
            categoria=categoria.upper(),
            localizacao=localizacao.upper(),
            descricao=descricao.upper()
        )
        
        if sucesso:
            QMessageBox.information(self, "Sucesso", "Produto atualizado com sucesso!")
            self.parent.atualizar_lista_produtos()
            self.accept()
        else:
            QMessageBox.critical(self, "Erro", "Erro ao atualizar produto!")


class InterfaceDeposito(QMainWindow):
    def __init__(self):
        super().__init__()
        self.deposito = GerenciadorDeposito()
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("Sistema de Gerenciamento de Depósito")
        self.setGeometry(100, 100, 1200, 700)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout()
        central_widget.setLayout(layout)
        
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)
        
        self.criar_aba_produtos()
        self.criar_aba_movimentacoes()
        self.criar_aba_relatorios()
        self.criar_aba_manutencao()
        self.criar_aba_sobre()
    
    def criar_aba_produtos(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        form_group = QGroupBox("Cadastro de Produto")
        form_layout = QGridLayout()
        
        form_layout.addWidget(QLabel("Nome:"), 0, 0)
        self.nome_input = QLineEdit()
        form_layout.addWidget(self.nome_input, 0, 1)
        
        form_layout.addWidget(QLabel("Categoria:"), 0, 2)
        self.categoria_input = QLineEdit()
        form_layout.addWidget(self.categoria_input, 0, 3)
        
        form_layout.addWidget(QLabel("Quantidade:"), 1, 0)
        self.quantidade_input = QLineEdit()
        form_layout.addWidget(self.quantidade_input, 1, 1)
        
        form_layout.addWidget(QLabel("Localização:"), 1, 2)
        self.localizacao_input = QLineEdit()
        form_layout.addWidget(self.localizacao_input, 1, 3)
        
        form_layout.addWidget(QLabel("Descrição:"), 2, 0)
        self.descricao_input = QLineEdit()
        form_layout.addWidget(self.descricao_input, 2, 1, 1, 3)
        
        form_group.setLayout(form_layout)
        layout.addWidget(form_group)
        
        btn_layout = QHBoxLayout()
        btn_adicionar = QPushButton("Adicionar Produto")
        btn_adicionar.clicked.connect(self.adicionar_produto)
        btn_layout.addWidget(btn_adicionar)
        
        btn_limpar = QPushButton("Limpar Campos")
        btn_limpar.clicked.connect(self.limpar_campos_produto)
        btn_layout.addWidget(btn_limpar)
        
        btn_atualizar = QPushButton("Atualizar Lista")
        btn_atualizar.clicked.connect(self.atualizar_lista_produtos)
        btn_layout.addWidget(btn_atualizar)
        
        layout.addLayout(btn_layout)
        
        busca_group = QGroupBox("Localizar Produto")
        busca_layout = QHBoxLayout()
        
        busca_layout.addWidget(QLabel("Buscar por nome:"))
        self.busca_input = QLineEdit()
        busca_layout.addWidget(self.busca_input)
        
        btn_buscar = QPushButton("Buscar")
        btn_buscar.clicked.connect(self.buscar_produtos)
        busca_layout.addWidget(btn_buscar)
        
        btn_limpar_busca = QPushButton("Limpar Busca")
        btn_limpar_busca.clicked.connect(self.limpar_busca)
        busca_layout.addWidget(btn_limpar_busca)
        
        busca_group.setLayout(busca_layout)
        layout.addWidget(busca_group)
        
        self.tabela_produtos = QTableWidget()
        self.tabela_produtos.setColumnCount(6)
        self.tabela_produtos.setHorizontalHeaderLabels(
            ['ID', 'Nome', 'Categoria', 'Quantidade', 'Localização', 'Descrição']
        )
        self.tabela_produtos.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.tabela_produtos)
        
        acoes_layout = QHBoxLayout()
        btn_entrada = QPushButton("Entrada")
        btn_entrada.clicked.connect(self.abrir_entrada)
        acoes_layout.addWidget(btn_entrada)
        
        btn_saida = QPushButton("Saida")
        btn_saida.clicked.connect(self.abrir_saida)
        acoes_layout.addWidget(btn_saida)
        
        btn_editar = QPushButton("Editar Produto")
        btn_editar.clicked.connect(self.editar_produto)
        acoes_layout.addWidget(btn_editar)
        
        layout.addLayout(acoes_layout)
        
        tab.setLayout(layout)
        self.tabs.addTab(tab, "Produtos")
        
        self.atualizar_lista_produtos()
    
    def criar_aba_movimentacoes(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        filtro_group = QGroupBox("Filtrar Movimentações")
        filtro_layout = QHBoxLayout()
        
        filtro_layout.addWidget(QLabel("Data Início:"))
        self.data_inicio = QDateEdit()
        self.data_inicio.setCalendarPopup(True)
        self.data_inicio.setDate(QDate.currentDate().addDays(-30))
        filtro_layout.addWidget(self.data_inicio)
        
        filtro_layout.addWidget(QLabel("Data Fim:"))
        self.data_fim = QDateEdit()
        self.data_fim.setCalendarPopup(True)
        self.data_fim.setDate(QDate.currentDate())
        filtro_layout.addWidget(self.data_fim)
        
        btn_filtrar = QPushButton("Filtrar")
        btn_filtrar.clicked.connect(self.filtrar_movimentacoes)
        filtro_layout.addWidget(btn_filtrar)
        
        btn_limpar_filtro = QPushButton("Limpar Filtro")
        btn_limpar_filtro.clicked.connect(self.limpar_filtro_movimentacoes)
        filtro_layout.addWidget(btn_limpar_filtro)
        
        filtro_group.setLayout(filtro_layout)
        layout.addWidget(filtro_group)
        
        self.tabela_movimentacoes = QTableWidget()
        self.tabela_movimentacoes.setColumnCount(6)
        self.tabela_movimentacoes.setHorizontalHeaderLabels(
            ['ID', 'Produto', 'Tipo', 'Quantidade', 'Data', 'Observação']
        )
        self.tabela_movimentacoes.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.tabela_movimentacoes)
        
        btn_atualizar = QPushButton("Atualizar")
        btn_atualizar.clicked.connect(self.atualizar_movimentacoes)
        layout.addWidget(btn_atualizar)
        
        tab.setLayout(layout)
        self.tabs.addTab(tab, "Movimentações")
        
        self.atualizar_movimentacoes()
    
    def criar_aba_relatorios(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Título
        titulo_label = QLabel("Relatórios do Sistema")
        titulo_label.setFont(QFont("Arial", 14, QFont.Bold))
        titulo_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(titulo_label)
        
        btn_layout = QHBoxLayout()
        
        btn_geral = QPushButton("Relatório Geral")
        btn_geral.clicked.connect(self.gerar_relatorio)
        btn_layout.addWidget(btn_geral)
        
        btn_baixo = QPushButton("Estoque Baixo")
        btn_baixo.clicked.connect(self.mostrar_estoque_baixo)
        btn_layout.addWidget(btn_baixo)
        
        btn_estoque = QPushButton("Produtos em Estoque")
        btn_estoque.clicked.connect(self.mostrar_produtos_em_estoque)
        btn_layout.addWidget(btn_estoque)
        
        btn_12meses = QPushButton("Movimentações 12 Meses")
        btn_12meses.clicked.connect(self.mostrar_movimentacoes_12_meses)
        btn_layout.addWidget(btn_12meses)
        
        layout.addLayout(btn_layout)
        
        # Botões de exportação
        export_layout = QHBoxLayout()
        
        btn_excel = QPushButton("Exportar para Excel")
        btn_excel.clicked.connect(self.exportar_para_excel)
        btn_excel.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; padding: 8px;")
        export_layout.addWidget(btn_excel)
        
        btn_pdf = QPushButton("Exportar para PDF")
        btn_pdf.clicked.connect(self.exportar_para_pdf)
        btn_pdf.setStyleSheet("background-color: #e74c3c; color: white; font-weight: bold; padding: 8px;")
        export_layout.addWidget(btn_pdf)
        
        layout.addLayout(export_layout)
        
        # Info do relatório
        self.info_relatorio = QLabel("")
        self.info_relatorio.setFont(QFont("Arial", 11, QFont.Bold))
        self.info_relatorio.setStyleSheet("color: #2c3e50; padding: 10px; background-color: #ecf0f1; border-radius: 5px;")
        layout.addWidget(self.info_relatorio)
        
        # Tabela de relatório
        self.tabela_relatorio = QTableWidget()
        self.tabela_relatorio.setAlternatingRowColors(True)
        self.tabela_relatorio.setStyleSheet("""
            QTableWidget {
                gridline-color: #bdc3c7;
                background-color: white;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 8px;
                font-weight: bold;
                border: 1px solid #2c3e50;
            }
        """)
        layout.addWidget(self.tabela_relatorio)
        
        tab.setLayout(layout)
        self.tabs.addTab(tab, "Relatórios")
    
    def criar_aba_manutencao(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        titulo = QLabel("Manutenção do Sistema")
        titulo.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(titulo)
        
        backup_group = QGroupBox("Backup do Banco de Dados")
        backup_layout = QVBoxLayout()
        
        desc = QLabel("Crie uma cópia de segurança do banco de dados.")
        backup_layout.addWidget(desc)
        
        btn_backup = QPushButton("Fazer Backup")
        btn_backup.clicked.connect(self.fazer_backup)
        backup_layout.addWidget(btn_backup)
        
        backup_group.setLayout(backup_layout)
        layout.addWidget(backup_group)
        
        restaurar_group = QGroupBox("Restaurar Banco de Dados")
        restaurar_layout = QVBoxLayout()
        
        desc2 = QLabel("Restaure o banco de dados a partir de um backup anterior.")
        restaurar_layout.addWidget(desc2)
        
        btn_restaurar = QPushButton("Restaurar Backup")
        btn_restaurar.clicked.connect(self.restaurar_backup)
        restaurar_layout.addWidget(btn_restaurar)
        
        restaurar_group.setLayout(restaurar_layout)
        layout.addWidget(restaurar_group)
        
        layout.addStretch()
        
        tab.setLayout(layout)
        self.tabs.addTab(tab, "Manutenção")
    
    def criar_aba_sobre(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        titulo = QLabel("Sistema de Gerenciamento de Depósito")
        titulo.setFont(QFont("Arial", 16, QFont.Bold))
        titulo.setAlignment(Qt.AlignCenter)
        layout.addWidget(titulo)
        
        versao = QLabel("Versão 1.2.0")
        versao.setAlignment(Qt.AlignCenter)
        layout.addWidget(versao)
        
        texto_sobre = QTextEdit()
        texto_sobre.setReadOnly(True)
        texto_sobre.setPlainText("""Sistema completo para controle e gerenciamento de estoque de depósitos.

PRINCIPAIS FUNCIONALIDADES:

- Cadastro de produtos com informações detalhadas
- Controle de entradas e saídas de mercadorias
- Busca e localização rápida de produtos
- Histórico completo de movimentações
- Filtros de movimentações por período
- Relatórios gerenciais em PDF e XLSX
- Backup e restauração do banco de dados

Desenvolvido por Felipe da Silva Braz
Licença GPLv3""")
        
        layout.addWidget(texto_sobre)
        
        tab.setLayout(layout)
        self.tabs.addTab(tab, "Sobre")
    
    def adicionar_produto(self):
        try:
            nome = self.nome_input.text().strip()
            categoria = self.categoria_input.text().strip()
            quantidade = int(self.quantidade_input.text() or 0)
            localizacao = self.localizacao_input.text().strip()
            descricao = self.descricao_input.text().strip()
            
            if not nome:
                QMessageBox.critical(self, "Erro", "Nome do produto é obrigatório!")
                return
            
            produto_id = self.deposito.adicionar_produto(
                nome=nome,
                quantidade=quantidade,
                descricao=descricao,
                categoria=categoria,
                localizacao=localizacao
            )
            
            QMessageBox.information(self, "Sucesso", f"Produto cadastrado com ID: {produto_id}")
            self.limpar_campos_produto()
            self.atualizar_lista_produtos()
            
        except ValueError:
            QMessageBox.critical(self, "Erro", "Digite um valor valido para quantidade!")
    
    def limpar_campos_produto(self):
        self.nome_input.clear()
        self.categoria_input.clear()
        self.quantidade_input.clear()
        self.localizacao_input.clear()
        self.descricao_input.clear()
    
    def atualizar_lista_produtos(self):
        produtos = self.deposito.listar_produtos()
        self.tabela_produtos.setRowCount(len(produtos))
        
        for i, p in enumerate(produtos):
            self.tabela_produtos.setItem(i, 0, QTableWidgetItem(str(p[0])))
            self.tabela_produtos.setItem(i, 1, QTableWidgetItem(p[1]))
            self.tabela_produtos.setItem(i, 2, QTableWidgetItem(p[3] or ""))
            self.tabela_produtos.setItem(i, 3, QTableWidgetItem(str(p[4])))
            self.tabela_produtos.setItem(i, 4, QTableWidgetItem(p[5] or ""))
            self.tabela_produtos.setItem(i, 5, QTableWidgetItem(p[2] or ""))
    
    def buscar_produtos(self):
        nome = self.busca_input.text().strip()
        
        if not nome:
            QMessageBox.warning(self, "Atenção", "Digite um nome para buscar!")
            return
        
        produtos = self.deposito.buscar_produto_por_nome(nome)
        
        if not produtos:
            QMessageBox.information(self, "Busca", "Nenhum produto encontrado!")
            return
        
        self.tabela_produtos.setRowCount(len(produtos))
        
        for i, p in enumerate(produtos):
            self.tabela_produtos.setItem(i, 0, QTableWidgetItem(str(p[0])))
            self.tabela_produtos.setItem(i, 1, QTableWidgetItem(p[1]))
            self.tabela_produtos.setItem(i, 2, QTableWidgetItem(p[3] or ""))
            self.tabela_produtos.setItem(i, 3, QTableWidgetItem(str(p[4])))
            self.tabela_produtos.setItem(i, 4, QTableWidgetItem(p[5] or ""))
            self.tabela_produtos.setItem(i, 5, QTableWidgetItem(p[2] or ""))
        
        QMessageBox.information(self, "Busca", f"{len(produtos)} produto(s) encontrado(s)!")
    
    def limpar_busca(self):
        self.busca_input.clear()
        self.atualizar_lista_produtos()
    
    def abrir_entrada(self):
        row = self.tabela_produtos.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Atenção", "Selecione um produto!")
            return
        
        produto_id = int(self.tabela_produtos.item(row, 0).text())
        produto_nome = self.tabela_produtos.item(row, 1).text()
        
        dialog = DialogMovimentacao(self, produto_id, produto_nome, "ENTRADA")
        dialog.exec_()
    
    def abrir_saida(self):
        row = self.tabela_produtos.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Atenção", "Selecione um produto!")
            return
        
        produto_id = int(self.tabela_produtos.item(row, 0).text())
        produto_nome = self.tabela_produtos.item(row, 1).text()
        quantidade_atual = int(self.tabela_produtos.item(row, 3).text())
        
        dialog = DialogMovimentacao(self, produto_id, produto_nome, "SAIDA", quantidade_atual)
        dialog.exec_()
    
    def editar_produto(self):
        row = self.tabela_produtos.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Atenção", "Selecione um produto!")
            return
        
        produto_id = int(self.tabela_produtos.item(row, 0).text())
        
        dialog = DialogEditarProduto(self, produto_id)
        dialog.exec_()
    
    def atualizar_movimentacoes(self):
        data_limite = (date.today() - timedelta(days=30)).strftime('%Y-%m-%d')
        movimentacoes = self.deposito.listar_movimentacoes(
            produto_id=None,
            data_inicio=data_limite,
            data_fim=None
        )
        
        self.tabela_movimentacoes.setRowCount(len(movimentacoes))
        
        for i, m in enumerate(movimentacoes):
            for j, valor in enumerate(m):
                self.tabela_movimentacoes.setItem(i, j, QTableWidgetItem(str(valor) if valor else ""))
    
    def filtrar_movimentacoes(self):
        data_inicio = self.data_inicio.date().toString("yyyy-MM-dd")
        data_fim = self.data_fim.date().toString("yyyy-MM-dd")
        
        movimentacoes = self.deposito.listar_movimentacoes(
            produto_id=None,
            data_inicio=data_inicio,
            data_fim=data_fim
        )
        
        if not movimentacoes:
            QMessageBox.information(self, "Filtro", "Nenhuma movimentação encontrada no período!")
            return
        
        self.tabela_movimentacoes.setRowCount(len(movimentacoes))
        
        for i, m in enumerate(movimentacoes):
            for j, valor in enumerate(m):
                self.tabela_movimentacoes.setItem(i, j, QTableWidgetItem(str(valor) if valor else ""))
        
        QMessageBox.information(self, "Filtro", f"{len(movimentacoes)} movimentação(ões) encontrada(s)!")
    
    def limpar_filtro_movimentacoes(self):
        self.data_inicio.setDate(QDate.currentDate().addDays(-30))
        self.data_fim.setDate(QDate.currentDate())
        self.atualizar_movimentacoes()
    
    def gerar_relatorio(self):
        relatorio = self.deposito.relatorio_estoque()
        
        # Configurar tabela
        self.tabela_relatorio.clear()
        self.tabela_relatorio.setColumnCount(2)
        self.tabela_relatorio.setHorizontalHeaderLabels(['METRICA', 'VALOR'])
        
        # Dados do resumo
        dados_resumo = [
            ['Total de Produtos Cadastrados', str(relatorio['total_produtos'])],
            ['Total de Itens em Estoque', str(relatorio['total_itens'])],
            ['', ''],  # Linha em branco
            ['CATEGORIA', 'PRODUTOS / ITENS'],
        ]
        
        # Adicionar categorias
        if relatorio['por_categoria']:
            for cat in relatorio['por_categoria']:
                categoria = cat[0] or "Sem categoria"
                dados_resumo.append([categoria, f"{cat[1]} produto(s) / {cat[2]} item(ns)"])
        else:
            dados_resumo.append(['Sem categorias', '-'])
        
        # Preencher tabela
        self.tabela_relatorio.setRowCount(len(dados_resumo))
        
        for i, linha in enumerate(dados_resumo):
            for j, valor in enumerate(linha):
                item = QTableWidgetItem(valor)
                
                # Destacar cabeçalhos
                if i == 0 or i == 1 or i == 3:
                    item.setFont(QFont("Arial", 10, QFont.Bold))
                    item.setBackground(Qt.lightGray)
                
                # Centralizar valores numéricos
                if j == 1 and i < 2:
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFont(QFont("Arial", 12, QFont.Bold))
                
                self.tabela_relatorio.setItem(i, j, item)
        
        self.tabela_relatorio.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tabela_relatorio.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        
        self.info_relatorio.setText(f"RELATÓRIO GERAL DO ESTOQUE - Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    
    def mostrar_estoque_baixo(self):
        produtos = self.deposito.produtos_estoque_baixo(10)
        
        # Configurar tabela
        self.tabela_relatorio.clear()
        self.tabela_relatorio.setColumnCount(5)
        self.tabela_relatorio.setHorizontalHeaderLabels(['ID', 'PRODUTO', 'CATEGORIA', 'QUANTIDADE', 'LOCALIZACAO'])
        
        if not produtos:
            self.tabela_relatorio.setRowCount(1)
            item = QTableWidgetItem("Nenhum produto com estoque baixo!")
            item.setTextAlignment(Qt.AlignCenter)
            item.setFont(QFont("Arial", 11, QFont.Bold))
            self.tabela_relatorio.setItem(0, 0, item)
            self.tabela_relatorio.setSpan(0, 0, 1, 5)
        else:
            self.tabela_relatorio.setRowCount(len(produtos))
            
            for i, p in enumerate(produtos):
                # ID
                item_id = QTableWidgetItem(str(p[0]))
                item_id.setTextAlignment(Qt.AlignCenter)
                self.tabela_relatorio.setItem(i, 0, item_id)
                
                # Nome
                self.tabela_relatorio.setItem(i, 1, QTableWidgetItem(p[1]))
                
                # Categoria
                self.tabela_relatorio.setItem(i, 2, QTableWidgetItem(p[2] or 'N/A'))
                
                # Quantidade (destacada em vermelho)
                item_qtd = QTableWidgetItem(str(p[3]))
                item_qtd.setTextAlignment(Qt.AlignCenter)
                item_qtd.setForeground(Qt.red)
                item_qtd.setFont(QFont("Arial", 10, QFont.Bold))
                self.tabela_relatorio.setItem(i, 3, item_qtd)
                
                # Localização
                self.tabela_relatorio.setItem(i, 4, QTableWidgetItem(p[4] or 'N/A'))
        
        # Configurar larguras dinâmicas das colunas
        header = self.tabela_relatorio.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID
        header.setSectionResizeMode(1, QHeaderView.Stretch)  # PRODUTO
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # CATEGORIA
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # QUANTIDADE
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # LOCALIZACAO
        
        self.info_relatorio.setText(f"PRODUTOS COM ESTOQUE BAIXO (<= 10) - Total: {len(produtos)} produto(s)")
    
    def mostrar_produtos_em_estoque(self):
        conn = self.deposito.conectar()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, nome, categoria, quantidade, localizacao
            FROM produtos
            WHERE quantidade > 0
            ORDER BY nome
        """)
        
        produtos = cursor.fetchall()
        conn.close()
        
        # Configurar tabela
        self.tabela_relatorio.clear()
        self.tabela_relatorio.setColumnCount(5)
        self.tabela_relatorio.setHorizontalHeaderLabels(['ID', 'PRODUTO', 'CATEGORIA', 'QUANTIDADE', 'LOCALIZACAO'])
        
        if not produtos:
            self.tabela_relatorio.setRowCount(1)
            item = QTableWidgetItem("Nenhum produto em estoque!")
            item.setTextAlignment(Qt.AlignCenter)
            item.setFont(QFont("Arial", 11, QFont.Bold))
            self.tabela_relatorio.setItem(0, 0, item)
            self.tabela_relatorio.setSpan(0, 0, 1, 5)
            total_itens = 0
        else:
            self.tabela_relatorio.setRowCount(len(produtos))
            total_itens = 0
            
            for i, p in enumerate(produtos):
                # ID
                item_id = QTableWidgetItem(str(p[0]))
                item_id.setTextAlignment(Qt.AlignCenter)
                self.tabela_relatorio.setItem(i, 0, item_id)
                
                # Nome
                self.tabela_relatorio.setItem(i, 1, QTableWidgetItem(p[1]))
                
                # Categoria
                self.tabela_relatorio.setItem(i, 2, QTableWidgetItem(p[2] or 'N/A'))
                
                # Quantidade (destacada em verde)
                item_qtd = QTableWidgetItem(str(p[3]))
                item_qtd.setTextAlignment(Qt.AlignCenter)
                item_qtd.setForeground(Qt.darkGreen)
                item_qtd.setFont(QFont("Arial", 10, QFont.Bold))
                self.tabela_relatorio.setItem(i, 3, item_qtd)
                
                # Localização
                self.tabela_relatorio.setItem(i, 4, QTableWidgetItem(p[4] or 'N/A'))
                
                total_itens += p[3]
        
        # Configurar larguras dinâmicas das colunas
        header = self.tabela_relatorio.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID
        header.setSectionResizeMode(1, QHeaderView.Stretch)  # PRODUTO
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # CATEGORIA
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # QUANTIDADE
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # LOCALIZACAO
        
        self.info_relatorio.setText(f"PRODUTOS EM ESTOQUE - Total: {len(produtos)} produto(s) / {total_itens} item(ns)")
    
    def mostrar_movimentacoes_12_meses(self):
        conn = self.deposito.conectar()
        cursor = conn.cursor()
        
        data_limite = (date.today() - timedelta(days=365)).strftime('%Y-%m-%d')
        
        cursor.execute("""
            SELECT 
                p.id,
                p.nome,
                p.categoria,
                COALESCE(SUM(CASE WHEN m.tipo = 'ENTRADA' THEN m.quantidade ELSE 0 END), 0) as total_entradas,
                COALESCE(SUM(CASE WHEN m.tipo = 'SAIDA' THEN m.quantidade ELSE 0 END), 0) as total_saidas,
                p.quantidade as estoque_atual
            FROM produtos p
            INNER JOIN movimentacoes m ON p.id = m.produto_id
            WHERE DATE(m.data_movimentacao) >= ?
            GROUP BY p.id, p.nome, p.categoria, p.quantidade
            ORDER BY p.nome
        """, (data_limite,))
        
        produtos = cursor.fetchall()
        conn.close()
        
        # Configurar tabela
        self.tabela_relatorio.clear()
        self.tabela_relatorio.setColumnCount(6)
        self.tabela_relatorio.setHorizontalHeaderLabels(['ID', 'PRODUTO', 'CATEGORIA', 'ENTRADAS', 'SAIDAS', 'SALDO ATUAL'])
        
        if not produtos:
            self.tabela_relatorio.setRowCount(1)
            item = QTableWidgetItem("Nenhuma movimentacao nos ultimos 12 meses!")
            item.setTextAlignment(Qt.AlignCenter)
            item.setFont(QFont("Arial", 11, QFont.Bold))
            self.tabela_relatorio.setItem(0, 0, item)
            self.tabela_relatorio.setSpan(0, 0, 1, 6)
            total_entradas = total_saidas = 0
        else:
            self.tabela_relatorio.setRowCount(len(produtos))
            total_entradas = 0
            total_saidas = 0
            
            for i, p in enumerate(produtos):
                # ID
                item_id = QTableWidgetItem(str(p[0]))
                item_id.setTextAlignment(Qt.AlignCenter)
                self.tabela_relatorio.setItem(i, 0, item_id)
                
                # Nome
                self.tabela_relatorio.setItem(i, 1, QTableWidgetItem(p[1]))
                
                # Categoria
                self.tabela_relatorio.setItem(i, 2, QTableWidgetItem(p[2] or 'N/A'))
                
                # Entradas (verde)
                item_ent = QTableWidgetItem(str(p[3]))
                item_ent.setTextAlignment(Qt.AlignCenter)
                item_ent.setForeground(Qt.darkGreen)
                item_ent.setFont(QFont("Arial", 10, QFont.Bold))
                self.tabela_relatorio.setItem(i, 3, item_ent)
                
                # Saídas (vermelho)
                item_sai = QTableWidgetItem(str(p[4]))
                item_sai.setTextAlignment(Qt.AlignCenter)
                item_sai.setForeground(Qt.red)
                item_sai.setFont(QFont("Arial", 10, QFont.Bold))
                self.tabela_relatorio.setItem(i, 4, item_sai)
                
                # Saldo (azul)
                item_saldo = QTableWidgetItem(str(p[5]))
                item_saldo.setTextAlignment(Qt.AlignCenter)
                item_saldo.setForeground(Qt.blue)
                item_saldo.setFont(QFont("Arial", 10, QFont.Bold))
                self.tabela_relatorio.setItem(i, 5, item_saldo)
                
                total_entradas += p[3]
                total_saidas += p[4]
        
        # Configurar larguras dinâmicas das colunas
        header = self.tabela_relatorio.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID
        header.setSectionResizeMode(1, QHeaderView.Stretch)  # PRODUTO
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # CATEGORIA
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # ENTRADAS
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # SAIDAS
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # SALDO ATUAL
        
        periodo = f"{data_limite} ate {date.today().strftime('%Y-%m-%d')}"
        self.info_relatorio.setText(
            f"MOVIMENTACOES DOS ULTIMOS 12 MESES ({periodo}) - "
            f"Produtos: {len(produtos)} | Entradas: {total_entradas} | Saidas: {total_saidas} | "
            f"Saldo: {total_entradas - total_saidas}"
        )
    
    def exportar_para_excel(self):
        if self.tabela_relatorio.rowCount() == 0:
            QMessageBox.warning(self, "Atenção", "Gere um relatório antes de exportar!")
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(
                self,
                "Erro",
                "Biblioteca openpyxl não instalada!\n\nInstale com: pip install openpyxl"
            )
            return
        
        arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar para Excel",
            f"relatorio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Arquivo Excel (*.xlsx)"
        )
        
        if arquivo:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Relatorio"
                
                # Estilos
                header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Alinhamento com quebra de linha
                wrap_alignment = Alignment(
                    horizontal='left',
                    vertical='top',
                    wrap_text=True
                )
                
                center_alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
                
                # Título do relatório
                ws.merge_cells('A1:' + chr(64 + self.tabela_relatorio.columnCount()) + '1')
                cell_titulo = ws['A1']
                cell_titulo.value = self.info_relatorio.text()
                cell_titulo.font = Font(bold=True, size=14)
                cell_titulo.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell_titulo.fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
                ws.row_dimensions[1].height = 30
                
                # Cabeçalhos
                for col in range(self.tabela_relatorio.columnCount()):
                    cell = ws.cell(row=2, column=col+1)
                    cell.value = self.tabela_relatorio.horizontalHeaderItem(col).text()
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                
                # Altura da linha de cabeçalho
                ws.row_dimensions[2].height = 25
                
                # Dados
                max_lengths = {}  # Para calcular largura das colunas
                
                for row in range(self.tabela_relatorio.rowCount()):
                    for col in range(self.tabela_relatorio.columnCount()):
                        item = self.tabela_relatorio.item(row, col)
                        if item:
                            cell = ws.cell(row=row+3, column=col+1)
                            cell.value = item.text()
                            cell.border = border
                            
                            # Aplicar alinhamento com quebra de linha
                            # Centralizar números, alinhar à esquerda textos
                            if col == 0 or (item.text().isdigit()):  # ID ou números
                                cell.alignment = center_alignment
                            else:
                                cell.alignment = wrap_alignment
                            
                            # Aplicar cores
                            if item.foreground().color().name() in ['#ff0000', '#ff0000ff']:  # Vermelho
                                cell.font = Font(bold=True, color="c0392b")
                            elif item.foreground().color().name() in ['#006400', '#006400ff']:  # Verde escuro
                                cell.font = Font(bold=True, color="27ae60")
                            elif item.foreground().color().name() in ['#0000ff', '#0000ffff']:  # Azul
                                cell.font = Font(bold=True, color="2980b9")
                            
                            # Calcular comprimento máximo para a coluna
                            text_length = len(str(item.text()))
                            if col not in max_lengths or text_length > max_lengths[col]:
                                max_lengths[col] = text_length
                            
                            # Definir altura da linha baseada no conteúdo
                            # Se o texto for muito longo, aumentar altura
                            if text_length > 50:
                                # Estimar altura necessária (aproximadamente 15 por linha)
                                estimated_lines = (text_length // 50) + 1
                                current_height = ws.row_dimensions[row+3].height or 15
                                new_height = max(current_height, estimated_lines * 15)
                                ws.row_dimensions[row+3].height = new_height
                
                # Ajustar largura das colunas baseado no conteúdo
                for col in range(1, self.tabela_relatorio.columnCount() + 1):
                    col_letter = chr(64 + col)
                    
                    # Calcular largura ideal
                    if (col - 1) in max_lengths:
                        # Limitar largura entre 15 e 60 caracteres
                        optimal_width = min(max(max_lengths[col - 1] + 2, 15), 60)
                    else:
                        optimal_width = 20
                    
                    ws.column_dimensions[col_letter].width = optimal_width
                
                # Congelar painéis (primeira linha e cabeçalho)
                ws.freeze_panes = 'A3'
                
                wb.save(arquivo)
                
                QMessageBox.information(self, "Sucesso", f"Relatório exportado para Excel!\n\n{arquivo}")
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao exportar para Excel:\n{str(e)}")
    
    def exportar_para_pdf(self):
        if self.tabela_relatorio.rowCount() == 0:
            QMessageBox.warning(self, "Atenção", "Gere um relatório antes de exportar!")
            return
        
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            QMessageBox.critical(
                self,
                "Erro",
                "Biblioteca ReportLab não instalada!\n\nInstale com: pip install reportlab"
            )
            return
        
        arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar para PDF",
            f"relatorio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            "Arquivo PDF (*.pdf)"
        )
        
        if arquivo:
            try:
                # Criar documento PDF com margens reduzidas
                doc = SimpleDocTemplate(
                    arquivo,
                    pagesize=landscape(A4),
                    rightMargin=20,
                    leftMargin=20,
                    topMargin=25,
                    bottomMargin=25
                )
                
                # Container para os elementos
                elements = []
                
                # Estilos
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=14,
                    textColor=colors.HexColor('#2c3e50'),
                    spaceAfter=8,
                    alignment=TA_CENTER,
                    fontName='Helvetica-Bold'
                )
                
                info_style = ParagraphStyle(
                    'CustomInfo',
                    parent=styles['Normal'],
                    fontSize=8,
                    textColor=colors.HexColor('#34495e'),
                    spaceAfter=10,
                    alignment=TA_CENTER,
                    fontName='Helvetica'
                )
                
                # Estilo para células com quebra de linha - COMPACTO
                cell_style_normal = ParagraphStyle(
                    'CellNormal',
                    parent=styles['Normal'],
                    fontSize=7,
                    leading=9,  # Espaçamento entre linhas reduzido
                    alignment=TA_LEFT,
                    fontName='Helvetica'
                )
                
                cell_style_bold = ParagraphStyle(
                    'CellBold',
                    parent=styles['Normal'],
                    fontSize=7,
                    leading=9,
                    alignment=TA_LEFT,
                    fontName='Helvetica-Bold'
                )
                
                cell_style_center = ParagraphStyle(
                    'CellCenter',
                    parent=styles['Normal'],
                    fontSize=7,
                    leading=9,
                    alignment=TA_CENTER,
                    fontName='Helvetica'
                )
                
                # Título
                titulo = Paragraph("Sistema de Gerenciamento de Depósito", title_style)
                elements.append(titulo)
                
                # Info do relatório
                info = Paragraph(self.info_relatorio.text(), info_style)
                elements.append(info)
                
                # Preparar dados da tabela
                data = []
                
                # Headers
                headers = []
                for col in range(self.tabela_relatorio.columnCount()):
                    header_item = self.tabela_relatorio.horizontalHeaderItem(col)
                    if header_item:
                        # Usar Paragraph para headers também
                        header_text = header_item.text()
                        header_para = Paragraph(f"<b>{header_text}</b>", cell_style_center)
                        headers.append(header_para)
                data.append(headers)
                
                # Dados - usar Paragraph para quebra automática
                for row in range(self.tabela_relatorio.rowCount()):
                    row_data = []
                    for col in range(self.tabela_relatorio.columnCount()):
                        item = self.tabela_relatorio.item(row, col)
                        if item:
                            text = item.text()
                            
                            # Verificar cor do texto para aplicar estilo
                            color = item.foreground().color()
                            
                            # Determinar estilo e cor
                            if color.name() in ['#ff0000', '#ff0000ff']:  # Vermelho
                                styled_text = f'<font color="red"><b>{text}</b></font>'
                                para = Paragraph(styled_text, cell_style_normal)
                            elif color.name() in ['#006400', '#006400ff']:  # Verde escuro
                                styled_text = f'<font color="green"><b>{text}</b></font>'
                                para = Paragraph(styled_text, cell_style_normal)
                            elif color.name() in ['#0000ff', '#0000ffff']:  # Azul
                                styled_text = f'<font color="blue"><b>{text}</b></font>'
                                para = Paragraph(styled_text, cell_style_normal)
                            else:
                                # Verificar se é número para centralizar
                                if col == 0 or text.isdigit():
                                    para = Paragraph(text, cell_style_center)
                                else:
                                    para = Paragraph(text, cell_style_normal)
                            
                            row_data.append(para)
                        else:
                            row_data.append(Paragraph("", cell_style_normal))
                    data.append(row_data)
                
                # Criar tabela PDF com larguras de coluna apropriadas
                # Calcular larguras baseadas no número de colunas
                num_cols = self.tabela_relatorio.columnCount()
                page_width = landscape(A4)[0] - 40  # Largura da página menos margens reduzidas
                
                # Distribuir larguras de forma inteligente
                if num_cols == 2:  # Relatório geral
                    col_widths = [page_width * 0.6, page_width * 0.4]
                elif num_cols == 5:  # Estoque baixo, produtos em estoque
                    col_widths = [
                        page_width * 0.06,  # ID
                        page_width * 0.38,  # Nome
                        page_width * 0.20,  # Categoria
                        page_width * 0.10,  # Quantidade
                        page_width * 0.26   # Localização
                    ]
                elif num_cols == 6:  # Movimentações 12 meses
                    col_widths = [
                        page_width * 0.06,  # ID
                        page_width * 0.36,  # Produto
                        page_width * 0.18,  # Categoria
                        page_width * 0.13,  # Entradas
                        page_width * 0.13,  # Saídas
                        page_width * 0.14   # Saldo
                    ]
                else:
                    # Dividir igualmente se não for um dos casos acima
                    col_widths = [page_width / num_cols] * num_cols
                
                pdf_table = Table(data, colWidths=col_widths, repeatRows=1)
                
                # Estilo da tabela - COMPACTO
                table_style = TableStyle([
                    # Header - padding reduzido
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('TOPPADDING', (0, 0), (-1, 0), 6),
                    
                    # Corpo da tabela - padding MUITO reduzido
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Alinhamento vertical no topo
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('TOPPADDING', (0, 1), (-1, -1), 3),     # Reduzido de 8 para 3
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 3),  # Reduzido de 8 para 3
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),    # Reduzido de 6 para 4
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),   # Reduzido de 6 para 4
                    
                    # Bordas mais finas
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    
                    # Linhas alternadas
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ])
                
                pdf_table.setStyle(table_style)
                elements.append(pdf_table)
                
                # Rodapé compacto
                elements.append(Spacer(1, 0.15 * inch))
                rodape_text = f"Gerado em: {datetime.now().strftime('%d/%m/%Y as %H:%M:%S')}"
                rodape_style = ParagraphStyle(
                    'Rodape',
                    parent=styles['Normal'],
                    fontSize=7,
                    textColor=colors.grey,
                    alignment=TA_CENTER
                )
                rodape = Paragraph(rodape_text, rodape_style)
                elements.append(rodape)
                
                # Gerar PDF
                doc.build(elements)
                
                QMessageBox.information(self, "Sucesso", f"Relatório exportado para PDF!\n\n{arquivo}")
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao exportar para PDF:\n{str(e)}")
    
    def fazer_backup(self):
        arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Backup",
            f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db",
            "Banco de Dados SQLite (*.db)"
        )
        
        if arquivo:
            try:
                shutil.copy2(self.deposito.db_name, arquivo)
                QMessageBox.information(self, "Sucesso", f"Backup realizado!\n\n{arquivo}")
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao fazer backup:\n{str(e)}")
    
    def restaurar_backup(self):
        resposta = QMessageBox.question(
            self,
            "Confirmação",
            "ATENÇÃO!\n\nEsta ação substituirá TODOS os dados atuais.\n\nDeseja continuar?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if resposta == QMessageBox.No:
            return
        
        arquivo, _ = QFileDialog.getOpenFileName(
            self,
            "Selecionar Backup",
            "",
            "Banco de Dados SQLite (*.db)"
        )
        
        if arquivo:
            try:
                self.deposito.conectar().close()
                shutil.copy2(arquivo, self.deposito.db_name)
                
                QMessageBox.information(self, "Sucesso", "Backup restaurado com sucesso!")
                
                self.atualizar_lista_produtos()
                self.atualizar_movimentacoes()
                
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao restaurar:\n{str(e)}")


def main():
    try:
        app = QApplication(sys.argv)
        window = InterfaceDeposito()
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        print(f"Erro ao iniciar aplicação: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
